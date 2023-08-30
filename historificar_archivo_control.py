import openpyxl
import pandas as pd
import os
import sys
from pydoc import stripid
import shutil
from contextlib import nullcontext

#Funcion: leer_parametros(archivo_parametros)
def leer_parametros(archivo_parametros):
    try:
        parametros = {}
        with open(archivo_parametros, "r") as archivo:
            for linea in archivo:
                etiqueta, valor = linea.strip().split("|")
                parametros[etiqueta] = valor
        return parametros
    except FileNotFoundError:
        print(f"leer_parametros: Error, El archivo {archivo_parametros} no se encontró.")
    except IOError:
        print(f"leer_parametros: Error, No se pudo leer el archivo {archivo_parametros}.")

#Funcion archive_backup(src_file, dst_file) el origen lo copia al dst incluyendo rutas
def archive_backup(src_file, dst_file):
    stError = False
    try:
        if not os.path.exists(src_file):
            print(f"archive_backup: Error, el fichero origen, no existe: {src_file}")
            stError = True
        else: 
            shutil.copy(src_file, dst_file)
            print("archive_backup: Backup realizado.")
            stError = False
    except Exception as e:
        print(f"archive_backup: How exceptional! {e}")
        print(f"archive_backup: Error, no se pudo realizar el backup del fichero destino")
        stError = True

    return stError        

#Funcion comparar_estados entre la pestaña Control (origen) y la pestaña (Historico) para historificar
def comparar_estados(archivo_origen, sh_origen, sh_destino, archivo_destino):
    try:
        # Abrir los archivos
        wblibro_origen = openpyxl.load_workbook(archivo_origen)
        wblibro_destino = openpyxl.load_workbook(archivo_destino)

        # Paso 1. Obtener la hoja de origen
        hoja_origen = wblibro_origen[sh_origen]
        # Ordenar la hoja de origen por la columna "Demanda"
        #hoja_origen.sort_by_column("Demandas", ascending=True)
        hoja_origen.sort_by_column(1, ascending=True)
        print(f'Punto de ordenación hoja origen pasado')        

        # Filtrar la hoja de origen
        #hoja_origen.auto_filter.filter_column("Estado", values=["Finalizada", "Dismissed"])
        #print(f'Punto de autofiltrado pasado')

        # Paso 2. Obtener la hoja de destino
        hoja_destino = wblibro_destino[sh_destino]
        #Ordenación por la columna "DEMANDA"
        hoja_destino.sort_by_column(1, ascending=True)
        print(f'Punto de ordenación hoja destino pasado')

        # Paso 3. Iterar sobre las filas de la hoja de origen
        #for fila_origen in hoja_origen.iter_rows():
        for fila_origen in hoja_origen.iter_rows(min_row=1, max_row=hoja_origen.max_row):
            # Obtener el valor de la columna "Demanda"
            demanda_origen = fila_origen[1].value
            # Obtener el valor de la columna "Estado" (columna 8)
            estado_origen = fila_origen[8].value

            if (estado_origen == "Finalizada") or (estado_origen == "Dismissed"):

                # Buscar la fila en la hoja de destino
                fila_destino = hoja_destino.find_row(demanda_origen, 0, 1)

                # Si la fila existe, comparar los estados
                if fila_destino is not None:
                    # Comprobar si el estado (columna 8) de la hoja de destino es el mismo que el de la hoja de origen
                    estado_destino = hoja_destino.cell(row=fila_destino, column=8).value

                    if estado_origen != estado_destino:
                        # Actualizar el estado de la hoja de destino
                        hoja_destino.cell(row=fila_destino, column=8).value = estado_origen
                        for i in range(2, hoja_origen.max_column + 1):
                            hoja_destino.cell(row=fila_destino, column=i).value = fila_origen[i].value
                else:
                    # La fila no existe, añadirla al final de la hoja de destino
                    hoja_destino.append(fila_origen)

        # Guardar el archivo de destino
        wblibro_destino.save(archivo_destino)
    except Exception as e:
        print(f"comparar_estados: How exceptional! {e}")    
    except:
        print("comparar_estados: Error inesperado:", sys.exc_info()[0])

#Funcion comparar_estados_historificacion
def comparar_estados_historificacion(archivo_origen, sh_origen, sh_destino, archivo_destino):
    try:
        stError = False
        # Abrir los archivos
        libro_origen = pd.read_excel(archivo_origen, sheet_name=sh_origen)
        libro_destino = pd.read_excel(archivo_destino, sheet_name=sh_destino)
        

        # Filtrar las filas de la hoja de origen
        #libro_origen = libro_origen[libro_origen["Estado"].isin(["Finalizada", "Dismissed"])]
        libro_origen = libro_origen.loc[libro_origen["Estado"].isin(["Finalizada", "Dismissed"])]
        
        # Ordenar las filas de la hoja de origen
        #libro_origen = libro_origen.sort_values("Demandas")
        libro_origen = libro_origen.sort_values(by=["Demandas"], ascending=True)

        # Obtener la lista de filas de la hoja de destino
        filas_destino = libro_destino.to_numpy()
        libro_destino = libro_destino.sort_values(by=["DEMANDA"], ascending=True)
        hoja_destino = libro_destino
        
        
        # Iterar sobre las filas de la hoja de origen
        for fila_origen in libro_origen.itertuples():
            # Obtener el valor de la columna "Demanda"        
            #demanda_origen = fila_origen["Demandas"]
            demanda_origen = fila_origen.Demandas
            print(f"Tratando fila demanda: {demanda_origen}")

            #estado_origen = fila_origen["Estado"]
            estado_origen = fila_origen.Estado
            print(f"Tratando fila estado: {estado_origen}")

            # Buscar la fila en la hoja de destino
            fila_destino = libro_destino.loc[libro_destino["DEMANDA"] == demanda_origen]
         
            # Si la fila existe, comparar los estados
            if not fila_destino.empty:
                # Actualizar el estado de la hoja de destino
                print(f'Encontrado - Fila destino demanda: {fila_destino.DEMANDA}')
                print(f'Encontrado - Fila destino estado: {fila_destino.Estado}')
                #print(f'Libro destino: {libro_destino["DEMANDA"]}')
                #libro_destino.loc[libro_destino["DEMANDA"] == demanda_origen, :] = fila_origen
                if fila_destino.size > 0:
                    #Actualizar camos de la fila destino
                    fila_destino.iloc[0] = fila_origen
                    print("Encontrado - Cambio de fila realizado")
            else:
                # La fila no existe, añadirla al final de la hoja de destino
                print("No Encontrado")
                print(f'No Encontrado - Fila destino demanda: {demanda_origen}')
                libro_destino = libro_destino.append(fila_origen)
                #libro_destino = libro_destino.append(fila_origen[0])
                #hoja_destino = hoja_destino.append(fila_origen)
                
                print("No Encontrado - Fila añadida")

        # Guardar el archivo de destino
        print("Llego hasta aqui (3)")
        libro_destino.to_excel(archivo_destino, index=False)
        return stError

    #except Exception as e:
    #    print(f"comparar_estados_historificacion: How exceptional! {e}")  
    #    stError = True
    #    return stError  
    except:
        print("comparar_estados_historificacion: Error inesperado:", sys.exc_info()[0])
        stError = True
        return stError  
    


def main():
    ruta_parametros = "D:\\IBD.GIT\\Python\\Automatizaciones\\src\\"
    archivo_parametros = "actualizar_filtro_parametros.txt"
    fparametros = ruta_parametros + archivo_parametros

    parametros = leer_parametros(fparametros)

    # Paso 1. Carga parametros fichero
    #Parametros leidos de fichero
    for ruta in parametros:
        if "RUTA_ARCHIVO_CONTROL" == ruta:
            ruta_archivo_control = str(parametros["RUTA_ARCHIVO_CONTROL"]).strip()
        if "RUTA_ARCHIVO_BACKUP" == ruta:
            ruta_archivo_backup = str(parametros["RUTA_ARCHIVO_BACKUP"]).strip()            
        if "ARCHIVO_CONTROL" == ruta:
            nombre_org_control = str(parametros["ARCHIVO_CONTROL"]).strip()
        if "PEST_CONTROL" == ruta:
            pest_control = str(parametros["PEST_CONTROL"]).strip()
        if "PEST_HISTORICO" == ruta:
            pest_historico = str(parametros["PEST_HISTORICO"]).strip()
    
    fOrig = nombre_org_control
    rOrig = ruta_archivo_control
    shOrig = pest_control
    shDest = pest_historico
    rBackup = ruta_archivo_backup
    fBackup = fOrig + '_v1'

    #1 paso) Backup del fichero filtro destino
    if rOrig is not nullcontext and fOrig is not nullcontext and rBackup is not nullcontext:
        stError = False
        print(" Fichero origen: ", fOrig)
        src_file = os.path.join(rOrig, fOrig)           #Fichero origen
        dst_file = os.path.join(rBackup, fBackup)       #Fichero backup
        if not os.path.exists(src_file):
            print(f"Error, el fichero origen no existe: {src_file}")
            resultado = "Error, el fichero origen no existe"
            stError = True     
    else:
        stError = True

    if fOrig is not nullcontext and rOrig is not nullcontext and shOrig is not nullcontext and shDest is not nullcontext and not stError:
        stError = archive_backup(src_file, dst_file)
        if stError:
            print(f'Error, el backup de archivos ha fallado.')
        else:
            #Comparar los estados
            #comparar_estados(src_file, shOrig, shDest, src_file)
            stError = comparar_estados_historificacion(src_file, shOrig, shDest, src_file)
            if not stError:
                print("historificar_archivo_control.py - Los estados se han comparado correctamente.")
            else:
                print("historificar_archivo_control.py - Error en la ejecución. ")



if __name__ == "__main__":
     # Only run the main function if this module is being run directly with `python main.py` or `python -m main`
    main()