# importing openpyxl module
import openpyxl
 
# Give the location of the file
My_path = "D:\\IBD.GIT\\Python\\Automatizaciones\\demo\\demo.xlsx"

# To open the workbook 
# workbook object is created
wb_obj = openpyxl.load_workbook(My_path)

# Get workbook active sheet object
# from the active attribute 
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj.cell(row = 1, column = 1)

# Loop will print all values
# of first column
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)


# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute.
sheet = wb.active

# Once have the Worksheet object,
# one can get its name from the
# title attribute.
sheet_title = sheet.title

print("active sheet title: " + sheet_title)

# Sheets can be added to workbook with the
# workbook object's create_sheet() method. 
wb.create_sheet(index = 1 , title = "demo sheet2")
  
wb.save(My_path)
