import os
from pydoc import stripid
import shutil
import sys
from typing import Self
from numpy import info
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from contextlib import nullcontext
import funcgeneral as fg


#UTILIDADES. Lectura de información de un fichero de parámetros
#Funcion: leer_parametros(archivo_parametros)
def leer_parametros(archivo_parametros):
    try:
        stError = False
        parametros = {}
        with open(archivo_parametros, "r") as archivo:
            for linea in archivo:
                etiqueta, valor = linea.strip().split("|")
                parametros[etiqueta] = valor
        return parametros
    except FileNotFoundError:
        stError = True
        print(f"leer_parametros: Error, El archivo {archivo_parametros} no se encontró.")

    except IOError:
        stError = True
        print(f"leer_parametros: Error, No se pudo leer el archivo {archivo_parametros}.")



#ESPECIFICAS
#Funcion: cargar_datos_desde_excel(origen, pestañas)
def cargar_datos_desde_excel(origen, pestañas):
    try:
        datos = {}
        for Mypestaña in pestañas:
            #Lectura archivo excel con un data frame (panda) según pestaña
            df = pd.read_excel(origen, sheet_name=Mypestaña)
            #devolvemos un vector con df por cada pestaña
            datos[Mypestaña] = df
        return datos
    except FileNotFoundError:
        print(f"cargar_datos_desde_excel: Error, El archivo {origen} no se encontró.")

    except IOError:
        print(f"cargar_datos_desde_excel: Error, No se pudo leer el archivo {origen}.")

    except:
        print("cargar_datos_en_excel: Error inesperado:", sys.exc_info()[0])


#Funcion: escribir_datos_en_excel(destino, datos)
#Informacion: datos es un array por pestaña
def escribir_datos_en_excel(destino, datos):
    try:
        Mypestaña = ""
        with pd.ExcelWriter(destino, engine='openpyxl') as writer:
            writer.book = load_workbook(destino)  #libro origen
            for Mypestaña, df in datos.items():
                df.to_excel(writer, sheet_name=Mypestaña, index=False)
            writer.save()
            writer.close()
    except FileNotFoundError:
        print(f"escribir_datos_en_excel: Error, El archivo {destino} no se encontró.")

    except IOError:
        print(f"escribir_datos_en_excel: Error, No se pudo leer el archivo {destino} ni pestaña {Mypestaña}.")

    except:
        print("escribir_datos_en_excel: Error inesperado:", sys.exc_info()[0])


# Funcion copiar_hoja(archivo_origen, pestaña_origen_pestaña_destino, archivo destino)
def copiar_hoja(archivo_origen, pestaña_origen, pestaña_destino, archivo_destino):
    stError = False
    try:    
        # Abrir los archivos, en dos workbook objects
        wblibro_origen = openpyxl.load_workbook(archivo_origen)
        wblibro_destino = openpyxl.load_workbook(archivo_destino)

        # Obtener la hoja de origen
        hoja_origen = wblibro_origen[pestaña_origen]

        # Obtener las dimensiones de la hoja de origen
        filas_origen = hoja_origen.max_row
        columnas_origen = hoja_origen.max_column

        # Obtener la hoja de destino
        hoja_destino = wblibro_destino[pestaña_destino]

        # Cell: the first row or column integer is 1
        # Eliminar los datos existentes de la hoja de destino
        for fila in range(1, filas_origen + 1):
            for columna in range(1, columnas_origen + 1):
                hoja_destino.cell(row=fila, column=columna).value = None

        # Copiar los datos de la hoja de origen a la hoja de destino
        for fila in range(1, filas_origen + 1):
            for columna in range(1, columnas_origen + 1):
                celda_origen = hoja_origen.cell(row=fila, column=columna)
                celda_destino = hoja_destino.cell(row=fila, column=columna)
                celda_destino.value = celda_origen.value

        # Guardar el archivo de destino
        wblibro_destino.save(archivo_destino)
    except:
        print("escribir_datos_en_excel: Error inesperado:", sys.exc_info()[0])

#Funcion archive_backup(src_file, dst_file) el origen lo copia al dst incluyendo rutas
def archive_backup(src_file, dst_file):
    stError = False
    try:
        if not os.path.exists(src_file):
            print(f"archive_backup: Error, el fichero origen, no existe: {src_file}")
            stError = True
        else: 
            shutil.copy(src_file, dst_file)
            print("archive_backup: Backup realizado.")
            stError = False
    except Exception as e:
        print(f"archive_backup: How exceptional! {e}")
        print(f"archive_backup: Error, no se pudo realizar el backup del fichero destino")
        stError = True

    return stError


#Funcion copiado hojas excel con backup
#   cp_Excel_bck(rOrig, fOrig, shOrig, rDest, fDest, shDest)
def cp_Excel_bck( mi_parametros ): 

  #inicializacion
  stError = False
  try: 

    #descarga del diccionario de parmetros
    rOrig = mi_parametros["rOrig"]
    fOrig = mi_parametros["fOrig"]
    shOrig = mi_parametros["shOrig"]
    rDest = mi_parametros["rDest"]
    fDest = mi_parametros["fDest"]
    shDest = mi_parametros["shDest"]
    rBackup = mi_parametros["rBackup"]

    resultado = ""
    fBackup = fDest + '_v1'
    #1 paso) Backup del fichero filtro destino
    if rDest is not nullcontext and fDest is not nullcontext:
        src_file = os.path.join(rDest, fDest)           #Fichero origen
        dst_file = os.path.join(rBackup, fBackup)       #Fichero destino
        
        #Backup archivo origen
        stError = archive_backup(src_file, dst_file)
    else:
        stError = True
   
    #2 paso) Obtener el fichero origen
    if fOrig is not nullcontext and rOrig is not nullcontext and shOrig is not nullcontext and not stError:
        print("cp_Excel_bck, Step1, Fichero origen: ", fOrig)
        src_file = os.path.join(rOrig, fOrig)
        if not os.path.exists(src_file):
            print(f"cp_Excel_bck: Error, el fichero origen no existe: {src_file}")
            resultado = "cp_Excel_bck: Error, el fichero origen no existe"
            stError = True
        else:
            # Load Excel File and give path to your file
            datosLec = {}
            datosLec = cargar_datos_desde_excel(src_file, shOrig)
            resultado = "cp_Excel_bck, Step1.2, Correcto, Carga archivo origen"
            stError = False
            print(resultado + ' ' + fOrig)

               
    else:
        resultado = "cp_Excel_bck: Error, Nombre vacio fichero origen/ruta origen/pestaña origen"
        print(resultado)
        stError = True

    #3 paso) Escritura del df en un archivo excel
    if fDest is not nullcontext and shDest is not nullcontext and not stError:
        print("cp_Excel_bck, Step2, tratamiento fichero destino: ", fDest)
        dst_file = os.path.join(rDest, fDest)
        if not os.path.exists(dst_file):
            print(f"cp_Excel_bck: Error, el fichero destino no existe: {src_file}")
            resultado = "cp_Excel_bck: Error, el fichero destino no existe"
            stError = True
        else:
            #Escribir en fichero destino
            #datosEsc = {}
            # datosEsc[shDest] = datosLec[shOrig]
            #escribir_datos_en_excel(dst_file, datosLec)
            copiar_hoja(src_file, shOrig, shDest, dst_file)
            print(f"cp_Excel_bck: Datos copiados con éxito en el archivo destino: {dst_file}")
            #df_Filtro.to_excel(dst_file, sheet_name = shDest, index=False)
            resultado = "cp_Excel_bck: Correcto, Escribe archivo destino"
            stError = False

    else:
        resultado = "cp_Excel_bckIncorrecto, hubo algun problema"
        print(resultado + ' ' + dst_file)
        stError = True
        
  except: 
      print("cp_Excel_bck: Error inesperado: ", sys.exc_info()[0])
      stError = True
  finally:
    #4 paso) Resultado
        return stError




def main():
    #CONFIGURACION GENERAL
    ruta_parametros = "D:\\IBD.GIT\\Python\\Automatizaciones\\src\\"
    archivo_parametros = "actualizar_filtro_parametros.txt"
    fparametros = ruta_parametros + archivo_parametros

    ruta_trazas_log = "D:\\IBD.GIT\\Python\\Automatizaciones\\log\\"
    archivo_trazas = "automatizacion_trazas.log"
    ftrazas = ruta_trazas_log + archivo_trazas

    #Activación de las trazas
    mytrazas = fg.Trazaslg(ftrazas)
    msg_trc = "actualizar_filtro_control.py"
    mytrazas.iniciar_traza(msg_trc)

    #Lectura del fichero de parámetros
    parametros = leer_parametros(fparametros)
    print('actualizar_filtro_control.py - En ejecución.')

    # Paso 1. Carga parametros fichero
    #Parametros leidos de fichero
    for ruta in parametros:
        if "RUTA_ARCHIVO_DESCARGAS" == ruta:
            ruta_archivo_descargas = str(parametros["RUTA_ARCHIVO_DESCARGAS"]).strip()
        if "RUTA_ARCHIVO_FILTROS" == ruta:
            ruta_archivo_filtros = str(parametros["RUTA_ARCHIVO_FILTROS"]).strip()
        if "RUTA_ARCHIVO_BACKUP" == ruta:
            ruta_archivo_backup = str(parametros["RUTA_ARCHIVO_BACKUP"]).strip()
        if "ARCHIVO_FILTRO" == ruta:
            nombre_dst_filtro = str(parametros["ARCHIVO_FILTRO"]).strip()
        if "ARCHIVO_TRELLO" == ruta:
            nombre_dst_trello = str(parametros["ARCHIVO_TRELLO"]).strip()
        if "PEST_FILTRO_ORIGEN" == ruta:
            pest_src_filtro = str(parametros["PEST_FILTRO_ORIGEN"]).strip()
        if "PEST_FILTRO_DESTINO" == ruta:
            pest_dst_filtro = str(parametros["PEST_FILTRO_DESTINO"]).strip()
        if "PEST_TRELLO_ORIGEN" == ruta:
            pest_src_trello = str(parametros["PEST_TRELLO_ORIGEN"]).strip()
        if "PEST_TRELLO_DESTINO" == ruta:
            pest_dst_trello = str(parametros["PEST_TRELLO_DESTINO"]).strip()

    # Paso 2.- Preparar la gestión del filtro
    #Paso 2.1. - Seleccionar el archivo del filtro
    fNombreOrigen = str(input("Introduzca nombre fichero origen (filtro): "))
    fSrcOrigen = ruta_archivo_descargas + fNombreOrigen + ".xlsx"
    src_hoja = pest_src_filtro.strip()
    dst_hoja = pest_dst_filtro.strip()
    stResultadoErr = False

    #lista de valores diccionario
    prm_cp_bck = {
        "rOrig": [],
        "fOrig": [],
        "shOrig": [],
        "rDest": [],
        "fDest": [],
        "shDest": [],
        "rBackup": [],
    }

    #2.2.- Informar el diccionario
    prm_cp_bck["rOrig"] = ruta_archivo_descargas    #parametro del fichero
    prm_cp_bck["fOrig"] = fSrcOrigen
    prm_cp_bck["shOrig"] = src_hoja
    prm_cp_bck["rDest"] = ruta_archivo_filtros      #parametro del fichero
    prm_cp_bck["fDest"] = nombre_dst_filtro         #parametro del fichero
    prm_cp_bck["shDest"] = dst_hoja
    prm_cp_bck["rBackup"] = ruta_archivo_backup


    msg_trc = "actualizar_filtro_control.py - El archivo " + archivo_parametros + " no se encontro."
    mytrazas.registrar_msg_traza(msg_trc)
    #Mostras valores del diccionario con el archivo de filtro
    #for llave in prm_cp_bck:
    #    #print("Llamada filtro: ", llave, ": ", prm_cp_bck[llave])
    #    print(f'actualizar_filtro_control.py - valores diccionario')
    #    print(f'Contenido diccionario {llave} : {prm_cp_bck[llave]}')

    #print(f"Elementos variables: ruta descargas: {ruta_archivo_descargas}; fOrig: {fSrcOrigen}; shOrig: {src_hoja}; rDest: {ruta_archivo_filtros}; fDest: {nombre_dst_filtro}; shDest: {dst_hoja}")
    #print("")
    
    #2.3.- copiar el archivo filtro original en backup.
    #stResultadoErr = cp_Excel_bck(rOrig = ruta_archivo_descargas, fOrig = fSrcOrigen, 
    #                        shOrig = src_hoja, rDest = ruta_archivo_filtros, 
    #                        fDest = nombre_dst_filtro, shDest = dst_hoja)
    stResultadoErr = cp_Excel_bck( prm_cp_bck)
    if not stResultadoErr:
        print(f'actualizar_filtro_control.py - Resultado de (cp_Excel_bck) copiar_planillas_backup: correcto.')
        msg_trc = "actualizar_filtro_control.py - Resultado de (cp_Excel_bck) copiar_planillas_backup: correcto."
        mytrazas.registrar_msg_traza(msg_trc)
    else:
        print(f'actualizar_filtro_control.py - Resultado de (cp_Excel_bck) copiar_planillas_backup: stResultadoErr = {stResultadoErr}')
        msg_err = "actualizar_filtro_control.py - Error (cp_Excel_bck) copiar_planillas_backup: stResultadoErr = " + str(stResultadoErr)
        mytrazas.registrar_err_traza(msg_err)

    #Paso 3.- Preparar la gestión del fichero Trello    
    stContinuar = str(input("Desea continuar (S/N): "))
    if (stContinuar.strip() == "S") or (stContinuar.strip() == "s"):

        fNombreOrigenTrello = str(input("Introduzca nombre fichero origen (trello): "))
        fSrcOrigen = ruta_archivo_descargas + fNombreOrigenTrello + ".xlsx"
        src_hoja = pest_src_trello.strip()
        dst_hoja = pest_dst_trello.strip()

        #Informar el diccionario de parametros
        prm_cp_bck["rOrig"] = ruta_archivo_descargas    #parametro del fichero
        prm_cp_bck["fOrig"] = fSrcOrigen
        prm_cp_bck["shOrig"] = src_hoja
        prm_cp_bck["rDest"] = ruta_archivo_filtros      #parametro del fichero
        prm_cp_bck["fDest"] = nombre_dst_trello         #parametro del fichero
        prm_cp_bck["shDest"] = dst_hoja
        prm_cp_bck["rBackup"] = ruta_archivo_backup

        #mostrar la información del diccionario con el archivo trello       
        #for llave in prm_cp_bck:
        #    print("Contenido dicc Trello: ", llave, ": ", prm_cp_bck[llave])

        stResultadoErr = False
        stResultadoErr = cp_Excel_bck(prm_cp_bck)
        print(f'actualizar_filtro_control.py - Actualización y copiado, Segundo archivo. stResultadoErrs = {stResultadoErr}')
        msg_trc = "actualizar_filtro_control.py - Actualización y copiado, Segundo archivo. stResultadoErr " + str(stResultadoErr)
        mytrazas.registrar_msg_traza(msg_trc)
        msg_trc = "actualizar_filtro_control.py - Fin de la ejecución."
        mytrazas.finalizar_traza(msg_trc) 
    else:
        print("Finalizado.")
        msg_trc = "actualizar_filtro_control.py - Fin de la ejecución."
        mytrazas.finalizar_traza(msg_trc)
        mytrazas.cerrar_traza()



if __name__ == "__main__":
     # Only run the main function if this module is being run directly with `python main.py` or `python -m main`
    main()
